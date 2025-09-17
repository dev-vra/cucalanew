# --- GERENCIADOR DE ESTOQUE CUCALA ---

import tkinter
from tkinter import ttk, filedialog, messagebox
import customtkinter as ctk
import pandas as pd
from datetime import datetime
import shutil
import os
import sys
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from pathlib import Path
import re
import json

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

# Configurações iniciais da aparência
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# --- ESTRUTURA RÍGIDA DE COLUNAS ---
COLUNAS_TOTAIS_ORDEM = [
    "UNIQUE ID", "CONT. REF", "REF.CUCALA", "NUMBER", "DATE", "SELLER", "BUYER", 
    "AGENT", "Nº AG", "Nº BUYER", "HVI", "ORIGIN", "GIN LOCATION", "FAZENDA(FARM NAME)", 
    "LOT NO.", "BALES", "Net Weight", "Tare", "GROSS KG", "GRADE", "STAPLE", "MIC.", 
    "GPT", "UHM", "LEAF", "COLOR", "CHARACTER", "TYPE", "P&D", "Lowest UHM", 
    "Bls<26,60UHM Bls<1,047 UHM", "Bls<27,39UHM Bls<1,079 UHM", 
    "Bls<28,18UHM Bls<1,110 UHM", "Bls<28,98UHM Bls<1,14 UHM", "Lowest Mic", 
    "Highest Mic", "Bls<3,3MIC", "Bls<3,4MIC", "Bls<3,5MIC", "Bls<3,6MIC", 
    "Bls<3,7MIC", "Bls>4,6MIC", "Bls>4,7MIC", "Bls>4,9MIC", "Bls>5 MIC", 
    "Lowest STR", "Bls<26gpt", "Bls<27gpt", "Blz<28gpt", "Blz<29gpt", 
    "Average - UNI", "Average - SFI", "Max- SFI", "Bls>10", "Bls>11", "Average - SCI", 
    "Min- SCI", "Bls<130", "Bls<140", "LEAF Min", "LEAF Max", "LEAF Average", 
    "Elong Min", "Elong Max", "Elong Average", "FDS<6 Elg", "Local", "Estado", 
    "Estoque", "Data Retirada", "Data devoluçao", "Data Descarte", "Motivo", "Operador"
]
COLUNAS_BASE = [
    "REF.CUCALA", "NUMBER", "SELLER", "BUYER", "ORIGIN", "GIN LOCATION", 
    "FAZENDA(FARM NAME)", "LOT NO.", "Local", "Estado", "Estoque", "Data Retirada", 
    "Data devoluçao", "Data Descarte"
]
COLUNAS_DETALHES = [
    "BALES", "Net Weight", "TYPE", "Lowest UHM", "Lowest Mic", "Lowest STR"
]
COLUNAS_MODIFICAVEIS = [
    "Local", "Estado", "Estoque", "Data Retirada", "Data devoluçao", "Data Descarte", "Motivo"
]

# --- MÓDULO DE GERENCIAMENTO DE CONFIGURAÇÃO E UTILIDADES ---
APP_NAME = "GerenciadorCucala"

def get_config_path():
    home_dir = Path.home()
    if os.name == 'nt': app_data_dir = home_dir / "AppData" / "Roaming" / APP_NAME
    elif os.uname().sysname == 'Darwin': app_data_dir = home_dir / "Library" / "Application Support" / APP_NAME
    else: app_data_dir = home_dir / ".config" / APP_NAME
    os.makedirs(app_data_dir, exist_ok=True)
    return app_data_dir / "config.json"

def save_settings(settings):
    with open(get_config_path(), 'w') as f: json.dump(settings, f, indent=4)

def load_settings():
    config_file = get_config_path()
    if not config_file.exists(): return {}
    try:
        with open(config_file, 'r') as f: return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError): return {}

def clean_string(text):
    if not isinstance(text, str): return text
    return re.sub(r'[^a-zA-Z0-9,./-]', '', text)

# --- JANELA DE LOGIN ---
class LoginWindow(ctk.CTk):
    def __init__(self):
        super().__init__()
        def load_users():
            with open(resource_path("data/users.json"), "r", encoding="utf-8") as f:
                all_users = json.load(f)
            return all_users.get("gerenciador", {})

        self.users = load_users()
        self.title("Login - Gerenciador de Estoque"); self.geometry("380x220"); self.resizable(False, False)
        self.grid_columnconfigure(0, weight=1); self.grid_rowconfigure(0, weight=1)
        main_frame = ctk.CTkFrame(self); main_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        main_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(main_frame, text="Acessar o Sistema", font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(10, 15))
        self.username_entry = ctk.CTkEntry(main_frame, placeholder_text="Usuário", width=250); self.username_entry.pack(pady=5, padx=10)
        self.password_entry = ctk.CTkEntry(main_frame, placeholder_text="Senha", show="*", width=250); self.password_entry.pack(pady=5, padx=10)
        self.password_entry.bind("<Return>", self._on_enter_key)
        login_button = ctk.CTkButton(main_frame, text="Entrar", command=self.check_login); login_button.pack(pady=(15, 10), padx=10)
    
    def _on_enter_key(self, event): self.check_login()
    
    def check_login(self):
        username = self.username_entry.get().lower()
        password = self.password_entry.get()
        if username in self.users and self.users[username] == password:
            self.destroy()
            StockManagerApp(username=username).mainloop()
        else:
            messagebox.showerror("Erro de Login", "Usuário ou senha inválidos.", parent=self); self.password_entry.delete(0, 'end')

# --- JANELAS DE FILTRO E AÇÕES ---
class ConsultaEstoqueWindow(ctk.CTkToplevel):
    def __init__(self, parent, dataframe):
        super().__init__(parent)
        self.transient(parent)
        self.title("Consulta de Estoque")
        self.geometry("800x600")
        self.grab_set()
        
        self.full_df = dataframe
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(expand=True, fill="both", padx=15, pady=15)

        # Frame para os filtros
        filter_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        filter_frame.pack(fill="x", pady=(0, 10))

        # Carrega os locais do JSON
        with open(resource_path("data/locais.json"), "r", encoding="utf-8") as f:
            self.locais_estoque = json.load(f)

        # --- Filtro Estoque ---
        # CORRIGIDO: Removi a primeira chamada duplicada do Label que causava o texto repetido.
        ctk.CTkLabel(filter_frame, text="Filtrar por Estoque:").pack(side="left")
        self.estoque_filter_var = ctk.StringVar(value="Todos")
        estoques = ["Todos"] + list(self.locais_estoque.keys())
        self.estoque_filter_menu = ctk.CTkOptionMenu(
            filter_frame, variable=self.estoque_filter_var,
            values=estoques, command=self._update_table
        )
        self.estoque_filter_menu.pack(side="left", padx=10)

        # --- Filtro Local ---
        ctk.CTkLabel(filter_frame, text="Filtrar por Local:").pack(side="left")
        self.local_filter_var = ctk.StringVar(value="Todos")
        # Criamos a lista de todos os locais únicos
        locais = ["Todos"] + sorted({loc for locs in self.locais_estoque.values() for loc in locs})
        self.local_filter_menu = ctk.CTkOptionMenu(
            filter_frame, variable=self.local_filter_var,
            values=locais, command=self._update_table
        )
        self.local_filter_menu.pack(side="left", padx=10)

        # ADICIONADO: Frame para conter a tabela (Treeview)
        tree_frame = ctk.CTkFrame(main_frame)
        tree_frame.pack(expand=True, fill="both", pady=(10, 0))

        # ADICIONADO: Chamadas para criar e popular a tabela na inicialização da janela
        self._setup_treeview(tree_frame)
        self._update_table()


    def _setup_treeview(self, parent_frame):
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background="#D3D3D3", foreground="black", rowheight=25, fieldbackground="#D3D3D3")
        style.map('Treeview', background=[('selected', '#347083')])
        
        self.tree = ttk.Treeview(parent_frame, style="Treeview")
        self.tree["columns"] = ("Local", "Total_Amostras", "Sellers", "Buyers")
        
        self.tree.column("#0", width=0, stretch=tkinter.NO)
        self.tree.heading("#0", text="", anchor=tkinter.W)
        
        self.tree.column("Local", anchor=tkinter.W, width=80)
        self.tree.heading("Local", text="Local", anchor=tkinter.W)
        
        self.tree.column("Total_Amostras", anchor=tkinter.CENTER, width=80)
        self.tree.heading("Total_Amostras", text="Total de Amostras", anchor=tkinter.CENTER)
        
        self.tree.column("Sellers", anchor=tkinter.W, width=250)
        self.tree.heading("Sellers", text="Sellers (Vendedores)", anchor=tkinter.W)

        self.tree.column("Buyers", anchor=tkinter.W, width=250)
        self.tree.heading("Buyers", text="Buyers (Compradores)", anchor=tkinter.W)
        
        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side='right', fill='y')
        hsb = ttk.Scrollbar(parent_frame, orient="horizontal", command=self.tree.xview)
        hsb.pack(side='bottom', fill='x')
        
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(expand=True, fill='both')

    def _update_table(self, filter_value=None):
        # Limpa a tabela antes de atualizar
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        # 1. Filtra apenas as amostras com estado 'Estocado'
        df_estocado = self.full_df[self.full_df['Estado'] == 'Estocado'].copy()
        if df_estocado.empty:
            return

        df_filtrado = df_estocado # Começa com todos os estocados

        # 2. Aplica o filtro de 'Estoque' (Estoque 1, Estoque 2, etc.)
        selected_estoque = self.estoque_filter_var.get()
        if selected_estoque != "Todos":
            df_filtrado = df_filtrado[df_filtrado['Estoque'] == selected_estoque]

        # 3. ADICIONADO: Aplica o filtro de 'Local' sobre o resultado anterior
        selected_local = self.local_filter_var.get()
        if selected_local != "Todos":
            df_filtrado = df_filtrado[df_filtrado['Local'] == selected_local]

        if df_filtrado.empty:
            return
        
        # Agrupa os resultados para a exibição
        summary = df_filtrado.groupby('Local').agg(
            Total_Amostras=('SELLER', 'size'),
            Sellers=('SELLER', lambda x: ', '.join(sorted(x.dropna().unique()))),
            Buyers=('BUYER', lambda x: ', '.join(sorted(x.dropna().unique())))
        ).reset_index()
        
        # Insere os dados na tabela
        for index, row in summary.iterrows():
            self.tree.insert("", "end", values=(row['Local'], row['Total_Amostras'], row['Sellers'], row['Buyers']))

class DescarteMassaWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent); self.transient(parent); self.title("Descarte em Massa"); self.geometry("600x500"); self.grab_set()
        self.parent_app = parent
        main_frame = ctk.CTkFrame(self); main_frame.pack(expand=True, fill="both", padx=15, pady=15)
        ctk.CTkLabel(main_frame, text="Cole os dados do Excel (duas colunas: Lote e Report/Number)").pack(anchor="w", pady=(0, 5))
        self.text_area = ctk.CTkTextbox(main_frame, height=300); self.text_area.pack(expand=True, fill="both")
        bottom_frame = ctk.CTkFrame(main_frame, fg_color="transparent"); bottom_frame.pack(fill="x", pady=(10, 0))
        ctk.CTkLabel(bottom_frame, text="Data do Descarte:").pack(side="left", padx=(0, 10))
        self.data_entry = ctk.CTkEntry(bottom_frame, placeholder_text="DD/MM/AAAA"); self.data_entry.pack(side="left")
        self.data_entry.insert(0, datetime.now().strftime("%d/%m/%Y"))
        ctk.CTkButton(bottom_frame, text="Descartar", command=self._on_descartar).pack(side="right")
        ctk.CTkButton(bottom_frame, text="Cancelar", fg_color="gray", command=self.destroy).pack(side="right", padx=10)

    def _on_descartar(self):
        texto_colado = self.text_area.get("1.0", "end-1c"); data_descarte = self.data_entry.get()
        if not texto_colado.strip(): messagebox.showwarning("Dados Vazios", "Por favor, cole os dados.", parent=self); return
        if not data_descarte: messagebox.showwarning("Data Inválida", "Por favor, informe a data.", parent=self); return
        items_para_descartar = []
        
        # Regex para encontrar o padrão de NUMBER (ex: ABC-123/1234)
        number_pattern = re.compile(r'([A-Z]+-\d+/\d{4})')
        
        for i, linha in enumerate(texto_colado.strip().splitlines()):
            if not linha.strip(): continue
            
            match = number_pattern.search(linha)
            
            if match:
                number = match.group(1)
                # Tudo antes do match é o lote bruto
                lote_bruto = linha[:match.start()].strip()
                # Processa o lote bruto para pegar o que vem antes do espaço
                lote_processado = lote_bruto.split(' ')[0]
                
                if lote_processado and number:
                    items_para_descartar.append((lote_processado, number, lote_bruto))
            else:
                messagebox.showwarning("Linha Inválida", f"A linha {i+1} ('{linha}') não contém um Report/Number válido e foi ignorada.", parent=self)
        
        if not items_para_descartar: messagebox.showerror("Nenhum Dado Válido", "Não foi possível processar nenhum par de Lote/Number.", parent=self); return
        self.parent_app.executar_descarte_massa(items_para_descartar, data_descarte, parent_window=self)
        self.destroy()

class AmostraFilterWindow(ctk.CTkToplevel):
    def __init__(self, parent, active_filters):
        super().__init__(parent); self.transient(parent); self.title("Filtrar Amostra (Valores Mínimos)"); self.geometry("400x300"); self.grab_set()
        self.parent_app = parent; self.filters = active_filters.get('gte', {}); self.entries = {}
        main_frame = ctk.CTkFrame(self); main_frame.pack(expand=True, fill="both", padx=20, pady=20)
        ctk.CTkLabel(main_frame, text="Digite os valores mínimos:", font=ctk.CTkFont(size=14)).grid(row=0, column=0, columnspan=2, pady=(0, 15), sticky="w")
        filter_fields = ['Lowest UHM', 'Lowest Mic', 'Lowest STR']
        for i, field in enumerate(filter_fields, start=1):
            ctk.CTkLabel(main_frame, text=f"{field}:").grid(row=i, column=0, padx=10, pady=10, sticky="e")
            entry = ctk.CTkEntry(main_frame, placeholder_text="Ex: 28.5"); entry.grid(row=i, column=1, padx=10, pady=10, sticky="ew")
            if field in self.filters: entry.insert(0, str(self.filters[field]))
            self.entries[field] = entry
        bottom_frame = ctk.CTkFrame(self, fg_color="transparent"); bottom_frame.pack(fill="x", padx=20, pady=(0, 10))
        ctk.CTkButton(bottom_frame, text="Aplicar Filtros", command=self._on_apply).pack(side="right")
        ctk.CTkButton(bottom_frame, text="Limpar e Fechar", fg_color="gray", command=self._on_clear).pack(side="right", padx=10)
    def _on_apply(self):
        new_gte_filters = {}
        for col, entry in self.entries.items():
            value = entry.get()
            if value:
                try: float(value.replace(',', '.')); new_gte_filters[col] = value
                except ValueError: messagebox.showerror("Valor Inválido", f"O valor '{value}' para '{col}' não é um número.", parent=self); return
        self.parent_app.execute_gte_filter(new_gte_filters, parent_window=self); self.destroy()
    def _on_clear(self): self.parent_app.execute_gte_filter({}); self.destroy()

class FilterChoiceWindow(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent); self.transient(parent); self.title("Tipo de Filtro"); self.geometry("350x150"); self.grab_set()
        self.parent_app = parent; main_frame = ctk.CTkFrame(self); main_frame.pack(expand=True, fill="both", padx=10, pady=10)
        ctk.CTkLabel(main_frame, text="Selecione o tipo de busca:").pack(pady=(5,15))
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent"); btn_frame.pack(fill="x", expand=True)
        ctk.CTkButton(btn_frame, text="Busca por Lote", command=self._on_busca_lote).pack(side="left", expand=True, padx=5)
        ctk.CTkButton(btn_frame, text="Filtrar Amostra", command=self._on_filtrar_amostra).pack(side="right", expand=True, padx=5)
    def _on_busca_lote(self): self.parent_app.open_filter_panel(mode="lote"); self.destroy()
    def _on_filtrar_amostra(self): self.parent_app.open_filter_panel(mode="amostra"); self.destroy()

class ValueSelectorWindow(ctk.CTkToplevel):
    def __init__(self, parent, column_name, unique_values, selected_values):
        super().__init__(parent); self.transient(parent); self.title(f"Selecionar valores para '{column_name}'"); self.geometry("400x500"); self.grab_set()
        self.parent = parent; self.column_name = column_name; self.all_values = sorted(unique_values); self.checkbox_vars = {}; self.checkbox_widgets = {}
        self.search_entry = ctk.CTkEntry(self, placeholder_text="Buscar valores..."); self.search_entry.pack(fill="x", padx=10, pady=(10, 5)); self.search_entry.bind("<KeyRelease>", self._filter_options)
        self.scrollable_frame = ctk.CTkScrollableFrame(self, label_text=f"Valores para {column_name}"); self.scrollable_frame.pack(expand=True, fill="both", padx=10, pady=5)
        for value in self.all_values:
            var = ctk.StringVar(value="on" if value in selected_values else "off")
            cb = ctk.CTkCheckBox(self.scrollable_frame, text=str(value), variable=var, onvalue="on", offvalue="off"); cb.pack(anchor="w", padx=10, pady=2)
            self.checkbox_vars[value] = var; self.checkbox_widgets[value] = cb
        btn_frame = ctk.CTkFrame(self, fg_color="transparent"); btn_frame.pack(fill="x", padx=10, pady=10)
        ctk.CTkButton(btn_frame, text="Confirmar", command=self._on_confirm).pack(side="right")
        ctk.CTkButton(btn_frame, text="Cancelar", fg_color="gray", command=self.destroy).pack(side="right", padx=10)
    def _filter_options(self, event=None):
        search_term = self.search_entry.get().lower()
        for value, widget in self.checkbox_widgets.items():
            if search_term in str(value).lower(): widget.pack(anchor="w", padx=10, pady=2)
            else: widget.pack_forget()
    def _on_confirm(self):
        selected = [value for value, var in self.checkbox_vars.items() if var.get() == "on"]
        self.parent.update_filter(self.column_name, selected); self.destroy()

class FilterWindow(ctk.CTkToplevel):
    def __init__(self, parent, dataframe, active_filters, filterable_columns):
        super().__init__(parent); self.transient(parent); self.title("Painel de Filtros"); self.geometry("600x400"); self.grab_set()
        self.parent_app = parent; self.df = dataframe; self.filters = active_filters.get('isin', {}); self.column_buttons = {}; self.filterable_columns = filterable_columns
        main_frame = ctk.CTkFrame(self); main_frame.pack(expand=True, fill="both", padx=10, pady=10)
        ctk.CTkLabel(main_frame, text="Clique na coluna para selecionar os valores:", font=ctk.CTkFont(size=14)).pack(anchor="w", pady=(0, 10))
        buttons_frame = ctk.CTkFrame(main_frame); buttons_frame.pack(expand=True, fill="both")
        for i, col in enumerate(self.filterable_columns):
            if col in self.df.columns:
                btn = ctk.CTkButton(buttons_frame, text=col, command=lambda c=col: self.open_value_selector(c)); btn.grid(row=i // 3, column=i % 3, padx=5, pady=5, sticky="ew")
                self.column_buttons[col] = btn
        self.update_button_states()
        bottom_frame = ctk.CTkFrame(self, fg_color="transparent"); bottom_frame.pack(fill="x", padx=10, pady=10)
        ctk.CTkButton(bottom_frame, text="Aplicar Filtros", command=self._on_apply).pack(side="right")
        ctk.CTkButton(bottom_frame, text="Fechar", fg_color="gray", command=self.destroy).pack(side="right", padx=10)
    def open_value_selector(self, column_name):
        pre_filtered_df = self.df.copy()
        for col, values in self.filters.items():
            if col != column_name: pre_filtered_df = pre_filtered_df[pre_filtered_df[col].isin(values)]
        unique_values = pre_filtered_df[column_name].dropna().unique()
        selected_values = self.filters.get(column_name, [])
        ValueSelectorWindow(self, column_name, unique_values, selected_values)
    def update_filter(self, column_name, selected_values):
        if selected_values: self.filters[column_name] = selected_values
        elif column_name in self.filters: del self.filters[column_name]
        self.update_button_states()
    def update_button_states(self):
        for col, btn in self.column_buttons.items():
            if col in self.filters: btn.configure(fg_color="#1F6AA5", text=f"{col} ({len(self.filters[col])})")
            else: btn.configure(fg_color=ctk.ThemeManager.theme["CTkButton"]["fg_color"], text=col)
    def _on_apply(self): self.parent_app.execute_isin_filter(self.filters, parent_window=self); self.destroy()

class MovementWindow(ctk.CTkToplevel):
    def __init__(self, parent, indices, rows_data):
        super().__init__(parent)
        self.transient(parent)
        self.title("Movimentar Lote(s)")
        self.geometry("500x450")
        self.grab_set()

        self.parent_app = parent
        self.indices = indices              # lista de índices selecionados
        self.rows_data = rows_data          # lista de dicts com os dados das linhas

        with open(resource_path("data/locais.json"), "r", encoding="utf-8") as f:
            self.locais_estoque = json.load(f)

        self._setup_widgets()
        self._load_initial_data()

    def _setup_widgets(self):
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(expand=True, fill="both", padx=20, pady=20)

        # título mostra quantos lotes foram selecionados
        ctk.CTkLabel(main_frame, text=f"{len(self.indices)} lote(s) selecionado(s)",
                     font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, columnspan=2, pady=10)

        ctk.CTkLabel(main_frame, text="Estado do Lote:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.estado_var = ctk.StringVar()
        self.estado_menu = ctk.CTkOptionMenu(
            main_frame, variable=self.estado_var,
            values=["", "Estocado", "Em uso", "Devolvido", "Descarte"],
            command=self._on_estado_change
        )
        self.estado_menu.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        self.conditional_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        self.conditional_frame.grid(row=2, column=0, columnspan=2, sticky="nsew")

        ctk.CTkButton(main_frame, text="Confirmar", command=self._on_confirm).grid(row=3, column=1, padx=10, pady=20, sticky="e")
        ctk.CTkButton(main_frame, text="Cancelar", fg_color="gray", command=self.destroy).grid(row=3, column=0, padx=10, pady=20, sticky="w")

    def _load_initial_data(self):
        # se vários lotes foram selecionados, o estado começa vazio
        if len(self.rows_data) == 1:
            initial_estado = self.rows_data[0].get("Estado", "")
            self.estado_var.set(initial_estado if pd.notna(initial_estado) else "")
        else:
            self.estado_var.set("")
        self._on_estado_change(self.estado_var.get())

    def _on_estado_change(self, selected_estado):
        for widget in self.conditional_frame.winfo_children():
            widget.destroy()

        if selected_estado == "Estocado":
            self._create_estocado_fields()
        elif selected_estado == "Em uso":
            self._create_em_uso_fields()
        elif selected_estado == "Devolvido":
            self._create_devolvido_fields()
        elif selected_estado == "Descarte":
            self._create_descarte_fields()

    def _create_estocado_fields(self):
        ctk.CTkLabel(self.conditional_frame, text="Estoque:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.estoque_var = ctk.StringVar()
        self.estoque_menu = ctk.CTkOptionMenu(
            self.conditional_frame, variable=self.estoque_var,
            values=["", "Estoque 1", "Estoque 2"], command=self._on_estoque_change
        )
        self.estoque_menu.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(self.conditional_frame, text="Local:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.local_var = ctk.StringVar()
        self.local_menu = ctk.CTkOptionMenu(self.conditional_frame, variable=self.local_var, values=[""])
        self.local_menu.grid(row=1, column=1, padx=10, pady=10, sticky="ew")
        self.local_menu.configure(state="disabled")

    def _on_estoque_change(self, selected_estoque):
        if selected_estoque in self.locais_estoque:
            self.local_menu.configure(values=self.locais_estoque[selected_estoque], state="normal")
        else:
            self.local_menu.configure(values=[""], state="disabled")
            self.local_var.set("")

    def _create_em_uso_fields(self):
        ctk.CTkLabel(self.conditional_frame, text="Data de Retirada:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.data_retirada_entry = ctk.CTkEntry(self.conditional_frame, placeholder_text="DD/MM/AAAA")
        self.data_retirada_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(self.conditional_frame, text="Motivo:").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.motivo_entry = ctk.CTkEntry(self.conditional_frame, placeholder_text="Ex: Análise, Amostra para cliente...")
        self.motivo_entry.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

    def _create_devolvido_fields(self):
        ctk.CTkLabel(self.conditional_frame, text="Data Devolução:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.data_devolucao_entry = ctk.CTkEntry(self.conditional_frame, placeholder_text="DD/MM/AAAA")
        self.data_devolucao_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

    def _create_descarte_fields(self):
        ctk.CTkLabel(self.conditional_frame, text="Data Descarte:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.data_descarte_entry = ctk.CTkEntry(self.conditional_frame, placeholder_text="DD/MM/AAAA")
        self.data_descarte_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

    def _on_confirm(self):
        updated_values = {col: '' for col in COLUNAS_MODIFICAVEIS}
        selected_estado = self.estado_var.get()
        updated_values['Estado'] = selected_estado
        updated_values['Operador'] = self.parent_app.current_user

        if selected_estado == "Estocado":
            estoque = self.estoque_var.get()
            local = self.local_var.get()
            if not estoque or not local:
                messagebox.showerror("Erro", "Campos 'Estoque' e 'Local' são obrigatórios.", parent=self)
                return
            updated_values['Estoque'] = estoque
            updated_values['Local'] = local
        elif selected_estado == "Em uso":
            updated_values['Data Retirada'] = self.data_retirada_entry.get()
            updated_values['Motivo'] = self.motivo_entry.get()
        elif selected_estado == "Devolvido":
            updated_values['Data devoluçao'] = self.data_devolucao_entry.get()
        elif selected_estado == "Descarte":
            updated_values['Data Descarte'] = self.data_descarte_entry.get()

        # aplica as alterações em todos os lotes selecionados
        for idx in self.indices:
            self.parent_app.update_row_data(idx, updated_values)

        self.destroy()

# --- JANELA PRINCIPAL DA APLICAÇÃO ---
class StockManagerApp(ctk.CTk):
    def __init__(self, username):
        super().__init__()
        self.current_user = username
        self.title("Gerenciador de Estoque CUCALA v1.0")
        self.geometry("1440x780")
        self.df_original = None
        self.active_filters = {'isin': {}, 'gte': {}}
        self.planilha_path = None
        self.details_switch_var = ctk.StringVar(value="off")
        self._setup_ui()
        self._auto_load_spreadsheet()

    def _setup_ui(self):
        top_frame = ctk.CTkFrame(self, height=50); top_frame.pack(side="top", fill="x", padx=10, pady=10)
        action_frame = ctk.CTkFrame(top_frame, fg_color="transparent"); action_frame.pack(side="left")
        self.btn_config = ctk.CTkButton(action_frame, text="Definir Planilha", command=self.select_and_set_spreadsheet_path); self.btn_config.pack(side="left", padx=5, pady=10)
        self.btn_filtrar = ctk.CTkButton(action_frame, text="Filtrar", state="disabled", command=self.aplicar_filtro); self.btn_filtrar.pack(side="left", padx=5, pady=10)
        self.btn_limpar_filtro = ctk.CTkButton(action_frame, text="Limpar Filtros", state="disabled", command=self.limpar_filtro); self.btn_limpar_filtro.pack(side="left", padx=5, pady=10)
        self.btn_movimentar = ctk.CTkButton(action_frame, text="Movimentar Lote", state="disabled", command=self.abrir_janela_movimentacao); self.btn_movimentar.pack(side="left", padx=(20, 5), pady=10)
        self.btn_descarte_massa = ctk.CTkButton(action_frame, text="Descarte em Massa", state="disabled", command=self.abrir_janela_descarte_massa); self.btn_descarte_massa.pack(side="left", padx=5, pady=10)
        self.btn_consultar_estoque = ctk.CTkButton(action_frame, text="Consultar Estoque", state="disabled", command=self.abrir_janela_consulta_estoque); self.btn_consultar_estoque.pack(side="left", padx=5, pady=10)
        view_frame = ctk.CTkFrame(top_frame, fg_color="transparent"); view_frame.pack(side="right")
        self.details_switch = ctk.CTkSwitch(view_frame, text="Exibir Detalhes", variable=self.details_switch_var, onvalue="on", offvalue="off", command=self._setup_and_populate_table); self.details_switch.pack(side="left", padx=10, pady=10)
        self.lbl_status_planilha = ctk.CTkLabel(view_frame, text="Nenhuma planilha carregada.", text_color="gray"); self.lbl_status_planilha.pack(side="left", padx=10)
        self.table_frame = ctk.CTkFrame(self); self.table_frame.pack(expand=True, fill="both", padx=10, pady=5)
        self._create_table_widget()
        bottom_frame = ctk.CTkFrame(self, height=50); bottom_frame.pack(side="bottom", fill="x", padx=10, pady=10)
        self.btn_salvar = ctk.CTkButton(bottom_frame, text="Salvar Alterações", state="disabled", command=self.salvar_alteracoes); self.btn_salvar.pack(side="right", padx=5, pady=10)

    def _auto_load_spreadsheet(self):
        settings = load_settings(); path = settings.get("database_path")
        if path and os.path.exists(path): self._load_spreadsheet_from_path(path)
        else: messagebox.showinfo("Bem-vindo!", "Nenhuma planilha definida. Clique em 'Definir Planilha' para começar.", parent=self)

    def select_and_set_spreadsheet_path(self):
        path = filedialog.askopenfilename(title="Selecione a Planilha Mestra", filetypes=[("Arquivos Excel", "*.xlsx")])
        if path and self._load_spreadsheet_from_path(path):
            save_settings({"database_path": path}); messagebox.showinfo("Sucesso", f"Caminho salvo! '{os.path.basename(path)}' será carregada automaticamente.", parent=self)

    def _load_spreadsheet_from_path(self, path):
        try:
            self.planilha_path = path
            df = pd.read_excel(self.planilha_path, sheet_name='database', dtype=str).dropna(how='all')
            # --- LIMPEZA APLICADA AQUI ---
            for col in ['LOT NO.', 'NUMBER', 'SELLER', 'BUYER']:
                if col in df.columns: df[col] = df[col].astype(str).apply(clean_string)
            self.df_original = df.fillna('')
            self.active_filters = {'isin': {}, 'gte': {}}
            self._setup_and_populate_table()
            for btn in [self.btn_filtrar, self.btn_limpar_filtro, self.btn_movimentar, self.btn_descarte_massa, self.btn_consultar_estoque, self.btn_salvar]: btn.configure(state="normal")
            self.lbl_status_planilha.configure(text=f"Carregada: {os.path.basename(self.planilha_path)}", text_color="green")
            return True
        except Exception as e:
            messagebox.showerror("Erro ao Ler Arquivo", f"Não foi possível ler a planilha em:\n{path}\n\nDetalhe: {e}", parent=self)
            self.lbl_status_planilha.configure(text="Falha ao carregar planilha.", text_color="red")
            return False

    def _create_table_widget(self):
        style = ttk.Style(); style.theme_use("default"); style.configure("Treeview", background="#D3D3D3", foreground="black", rowheight=25, fieldbackground="#D3D3D3"); style.map('Treeview', background=[('selected', '#347083')])
        self.tree = ttk.Treeview(self.table_frame, style="Treeview"); vsb = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview); vsb.pack(side='right', fill='y'); hsb = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.tree.xview); hsb.pack(side='bottom', fill='x')
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set); self.tree.pack(expand=True, fill='both')

    def _setup_and_populate_table(self):
        self.tree.delete(*self.tree.get_children())
        if self.details_switch_var.get() == "on": current_visible_columns = COLUNAS_BASE + COLUNAS_DETALHES
        else: current_visible_columns = COLUNAS_BASE
        self.df_display = self.df_original.reindex(columns=current_visible_columns)
        self.tree["columns"] = current_visible_columns
        self.tree.column("#0", width=0, stretch=tkinter.NO); self.tree.heading("#0", text="", anchor=tkinter.W)
        for col in current_visible_columns: self.tree.column(col, anchor=tkinter.W, width=120); self.tree.heading(col, text=col, anchor=tkinter.W)
        self.execute_filter()

    def _repopulate_table_rows(self, dataframe_to_show):
        self.tree.delete(*self.tree.get_children())
        df_view = dataframe_to_show.fillna('')
        for index, row in df_view.iterrows(): self.tree.insert("", "end", iid=index, values=list(row))

    def update_row_data(self, index, new_values):
        for col, value in new_values.items():
            if col not in self.df_original.columns: self.df_original[col] = ''
            self.df_original.loc[index, col] = value
        self._setup_and_populate_table()

    def aplicar_filtro(self):
        if self.df_original is None: return
        FilterChoiceWindow(self)

    def open_filter_panel(self, mode):
        if mode == "lote":
            cols = [
             'SELLER', 'BUYER', 'GIN LOCATION',
             'FAZENDA(FARM NAME)', 'LOT NO.',
             'NUMBER', 'REF.CUCALA'
            ]
            FilterWindow(self, self.df_original, self.active_filters, cols)
        elif mode == "amostra": AmostraFilterWindow(self, self.active_filters)
    
    def execute_isin_filter(self, filters, parent_window=None):
        self.active_filters['isin'] = filters; self.execute_filter(parent_window=parent_window)
    
    def execute_gte_filter(self, filters, parent_window=None):
        self.active_filters['gte'] = filters; self.execute_filter(parent_window=parent_window)
        
    def execute_filter(self, parent_window=None):
        df_to_show = self.df_display.copy()
        isin_filters = self.active_filters.get('isin', {}); gte_filters = self.active_filters.get('gte', {})
        for column, values in isin_filters.items():
            df_to_show = df_to_show[df_to_show[column].astype(str).isin([str(v) for v in values])]
        for column, min_value in gte_filters.items():
            df_to_show[column] = pd.to_numeric(df_to_show[column], errors='coerce')
            min_value_float = float(str(min_value).replace(',', '.'))
            df_to_show = df_to_show[df_to_show[column] >= min_value_float]
        self._repopulate_table_rows(df_to_show)
        parent = parent_window if parent_window else self
        if len(isin_filters) + len(gte_filters) > 0:
            messagebox.showinfo("Filtro Aplicado", f"{len(df_to_show)} registros encontrados.", parent=parent)
        
    def limpar_filtro(self):
        self.active_filters = {'isin': {}, 'gte': {}}; self._repopulate_table_rows(self.df_display)
        messagebox.showinfo("Filtro Removido", "A visualização da tabela foi restaurada.", parent=self)

    def abrir_janela_movimentacao(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showwarning("Nenhuma Seleção", "Por favor, selecione ao menos um lote.", parent=self)
            return

        indices = [int(i) for i in selected_items]
        rows_data = [self.df_original.loc[idx].to_dict() for idx in indices]

        MovementWindow(self, indices, rows_data)

    def abrir_janela_descarte_massa(self):
        if self.df_original is None: return
        DescarteMassaWindow(self)

    def abrir_janela_consulta_estoque(self):
        if self.df_original is None: return
        ConsultaEstoqueWindow(self, self.df_original)

    def executar_descarte_massa(self, items_para_descartar, data_descarte, parent_window=None):
        encontrados = 0; nao_encontrados = []
        df = self.df_original
        for lote_proc, number, lote_bruto in items_para_descartar:
            # Compara os dados já limpos
            condicao = (df['LOT NO.'] == lote_proc) & (df['NUMBER'] == number)
            indices = df[condicao].index
            if not indices.empty:
                df.loc[indices, 'Estado'] = 'Descarte'; df.loc[indices, 'Data Descarte'] = data_descarte
                df.loc[indices, 'Operador'] = self.current_user
                encontrados += len(indices)
            else:
                nao_encontrados.append(f"{lote_bruto} | {number}")
        self.df_original = df
        self._setup_and_populate_table()
        mensagem = f"{encontrados} lote(s) marcados para descarte.\n\n"
        if nao_encontrados: mensagem += f"{len(nao_encontrados)} lote(s) não encontrados:\n" + "\n".join(nao_encontrados)
        parent = parent_window if parent_window else self
        messagebox.showinfo("Processo Concluído", mensagem, parent=parent)

    def salvar_alteracoes(self):
        if not self.planilha_path or self.df_original is None: messagebox.showerror("Erro", "Nenhuma planilha carregada.", parent=self); return
        if not messagebox.askyesno("Confirmar", "Salvar todas as alterações? Um backup será criado.", parent=self): return
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S"); base, ext = os.path.splitext(self.planilha_path); backup_path = f"{base}_backup_{timestamp}{ext}"; shutil.copy(self.planilha_path, backup_path)
            df_to_save = self.df_original.copy()
            for col in COLUNAS_TOTAIS_ORDEM:
                if col not in df_to_save.columns: df_to_save[col] = ''
            df_to_save = df_to_save[COLUNAS_TOTAIS_ORDEM]
            workbook = load_workbook(self.planilha_path)
            if "database" in workbook.sheetnames: del workbook["database"]
            db_sheet = workbook.create_sheet("database")
            for r in dataframe_to_rows(df_to_save, index=False, header=True): db_sheet.append(r)
            main_sheet_name = workbook.sheetnames[0]
            if main_sheet_name != "database":
                main_sheet = workbook[main_sheet_name]
                if main_sheet.max_row > 1: main_sheet.delete_rows(2, main_sheet.max_row)
                for r_idx, row in enumerate(dataframe_to_rows(df_to_save, index=False, header=False), 2):
                    for c_idx, value in enumerate(row, 1): main_sheet.cell(row=r_idx, column=c_idx, value=value)
            workbook.save(self.planilha_path)
            messagebox.showinfo("Sucesso", f"Alterações salvas!\nBackup: {os.path.basename(backup_path)}", parent=self)
        except PermissionError: messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar.\nA planilha '{os.path.basename(self.planilha_path)}' pode estar aberta.", parent=self)
        except Exception as e: messagebox.showerror("Erro ao Salvar", f"Ocorreu um erro inesperado.\nDetalhe: {e}", parent=self)

# --- INICIALIZAÇÃO DA APLICAÇÃO ---
if __name__ == "__main__":
    LoginWindow().mainloop()