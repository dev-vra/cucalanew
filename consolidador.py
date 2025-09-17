# SEÇÃO DE IMPORTAÇÕES DE BIBLIOTECAS
import pandas as pd
import openpyxl
import os
import shutil
import sys
import tkinter
from tkinter import filedialog, messagebox
import logging
import traceback
import threading
import queue
from datetime import datetime
from dateutil.parser import parse
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import customtkinter as ctk
from PIL import Image
from pathlib import Path
import configparser
import json

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

# --- FUNÇÃO: Define o local correto para salvar dados no macOS ---
def get_app_support_dir():
    app_support_path = Path.home() / "Library" / "Application Support" / "ConsolidadorCUCALA"
    app_support_path.mkdir(parents=True, exist_ok=True)
    return app_support_path

# --- CONFIGURAÇÃO DO LOGGING E ARQUIVOS ---
APP_SUPPORT_DIR = get_app_support_dir()
LOG_FILE_PATH = APP_SUPPORT_DIR / "consolidador.log"
CONFIG_FILE_PATH = APP_SUPPORT_DIR / "config.ini" 

logging.basicConfig(filename=LOG_FILE_PATH, level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

# NÚCLEO LÓGICO
def processar_arquivo_origem(arquivo, colunas_mestra, username, log_queue):
    def log(msg):
        log_queue.put(msg)
    
    try:
        workbook = openpyxl.load_workbook(arquivo, data_only=True)
        sheet = workbook.worksheets[0]
        
        number_raw = sheet['C3'].value
        seller_raw = sheet['C8'].value
        
        number = str(number_raw).strip() if number_raw is not None else None
        seller = str(seller_raw).strip() if seller_raw is not None else None
        
        if not number or not seller:
            log(f"  -> ERRO: Não foi possível extrair NUMBER(C3) ou SELLER(C8) de {os.path.basename(arquivo)}.")
            return None, None

        ref_cucala, date_raw, buyer, agent, n_ag, n_buyer = (
            sheet['C13'].value, sheet['C4'].value, sheet['C9'].value,
            sheet['C10'].value, sheet['E9'].value, sheet['E10'].value
        )
        date_formatada = date_raw
        if date_raw:
            try: date_formatada = parse(str(date_raw)).strftime('%d/%m/%Y')
            except Exception: pass
        
        df_dados = pd.read_excel(arquivo, engine='openpyxl', skiprows=13)
        
        mapa_de_traducao = {
            'ORIGEM': 'ORIGIN', 'MUNICIPIO': 'GIN LOCATION', 'FAZENDA': 'FAZENDA(FARM NAME)',
            'FAZENDA (FARM NAME)': 'FAZENDA(FARM NAME)', 'LOTE': 'LOT NO.', 'FARDOS': 'BALES',
            'P.LIQUIDO': 'Net Weight', 'TARA': 'Tare', 'P.BRUTO': 'GROSS KG', 'TIPO': 'GRADE',
            'FIBRA': 'STAPLE', 'FOLHA': 'LEAF', 'COR': 'COLOR', 'BENEFICIO': 'CHARACTER',
            'TIPO DO VENDEDOR': 'TYPE', 'TYPE AGREED': 'TYPE', 'OBSERVAÇÃO': 'P&D'
        }
        df_dados.columns = df_dados.columns.str.strip()
        df_dados.rename(columns=mapa_de_traducao, inplace=True)

        if 'HVI' in df_dados.columns:
            indices_total = df_dados[df_dados['HVI'] == 'TOTAL'].index
            if not indices_total.empty: df_dados = df_dados.loc[:indices_total[0] - 1]
        
        df_dados = df_dados.dropna(axis=1, how='all')
        if not df_dados.empty: df_dados = df_dados.dropna(subset=[df_dados.columns[0]], how='all')
        if df_dados.empty: return None, None

        # --- INÍCIO DA NOVA ALTERAÇÃO ---
        # Função que converte para número se possível, senão mantém como texto
        def converter_valor(valor):
            if valor is None:
                return None
            try:
                # Tenta converter para float. Funciona para inteiros e decimais.
                return float(valor)
            except (ValueError, TypeError):
                # Se falhar, significa que contém texto (ex: 'R130'). Retorna o valor original.
                return valor

        colunas_a_tratar = ['BALES', 'GROSS KG']
        for col in colunas_a_tratar:
            if col in df_dados.columns:
                df_dados[col] = df_dados[col].apply(converter_valor)
        # --- FIM DA NOVA ALTERAÇÃO ---
        
        df_final = df_dados.copy()
        df_final.insert(0, 'CONT. REF', range(1, len(df_final) + 1))
        df_final.insert(0, 'Nº BUYER', n_buyer); df_final.insert(0, 'Nº AG', n_ag); df_final.insert(0, 'AGENT', agent)
        df_final.insert(0, 'BUYER', buyer); df_final.insert(0, 'SELLER', seller); df_final.insert(0, 'DATE', date_formatada)
        df_final.insert(0, 'NUMBER', number); df_final.insert(0, 'REF.CUCALA', ref_cucala)
        
        if username:
            df_final.insert(0, 'userid', username.capitalize())

        df_alinhado = df_final.reindex(columns=colunas_mestra)
        
        log(f"  -> Sucesso: {len(df_alinhado)} linhas extraídas.")
        return df_alinhado, (number, seller)
    except Exception as e:
        log(f"  -> ERRO GERAL ao processar o arquivo {os.path.basename(arquivo)}: {e}")
        log(traceback.format_exc())
        return None, None

def executar_logica_consolidacao(planilha_mestra, lista_arquivos_origem, log_queue, username, modo_atualizacao=False):
    def log(msg):
        log_queue.put(msg)
    try:
        log("Iniciando processo...")
        try:
            log("Lendo dados existentes da aba 'database'...")
            df_mestra = pd.read_excel(planilha_mestra, sheet_name='database', header=0)
            colunas_mestra_originais = df_mestra.columns.tolist()
        except Exception:
            log("AVISO: Aba 'database' não encontrada ou vazia. Assumindo uma base de dados nova.")
            try:
                df_mestra_estrutura = pd.read_excel(planilha_mestra, header=1)
                colunas_mestra_originais = df_mestra_estrutura.columns.tolist()
                df_mestra = pd.DataFrame(columns=colunas_mestra_originais)
            except Exception as e:
                log(f"ERRO: Não foi possível definir a estrutura de colunas a partir do arquivo mestre. Verifique o arquivo. Erro: {e}")
                return False, "Falha ao ler a estrutura do arquivo mestre."

        df_mestra_limpa = df_mestra.dropna(subset=['CONT. REF'], how='all').copy() if not df_mestra.empty else df_mestra

    except Exception as e:
        log(f"ERRO: Não foi possível ler a planilha mestra: {e}")
        return False, "Falha ao ler planilha mestra."

    arquivos_processados = 0
    df_final = df_mestra_limpa.copy()

    if modo_atualizacao:
        log("--- Modo de Atualização Ativado ---")
        if not df_final.empty:
            df_final['NUMBER'] = df_final['NUMBER'].astype(str).str.strip()
            df_final['SELLER'] = df_final['SELLER'].astype(str).str.strip()
            
        for arquivo in lista_arquivos_origem:
            df_processado, id_lancamento = processar_arquivo_origem(arquivo, colunas_mestra_originais, username, log_queue)
            if df_processado is None: continue

            number, seller = id_lancamento
            
            indices_para_atualizar = df_final[(df_final['NUMBER'] == number) & (df_final['SELLER'] == seller)].index
            
            if not indices_para_atualizar.empty:
                log(f"Lançamento '{number} - {seller}' encontrado. Atualizando...")
                if len(indices_para_atualizar) != len(df_processado):
                    log(f"  -> ERRO: A quantidade de linhas no arquivo de atualização ({len(df_processado)}) é diferente da existente ({len(indices_para_atualizar)}). Atualização cancelada.")
                    continue
                
                df_processado.index = indices_para_atualizar
                df_final.update(df_processado)
                arquivos_processados += 1
            else:
                log(f"AVISO: Lançamento '{number} - {seller}' não encontrado para ser atualizado.")
    else:
        log("--- Modo de Adição de Novos Lançamentos ---")
        dados_para_adicionar = []
        
        if not df_final.empty:
            df_final['NUMBER'] = df_final['NUMBER'].astype(str).str.strip()
            df_final['SELLER'] = df_final['SELLER'].astype(str).str.strip()

        for arquivo in lista_arquivos_origem:
            df_processado, id_lancamento = processar_arquivo_origem(arquivo, colunas_mestra_originais, username, log_queue)
            if df_processado is None: continue

            number, seller = id_lancamento
            
            existe = not df_final[(df_final['NUMBER'] == number) & (df_final['SELLER'] == seller)].empty

            if not existe:
                log(f"Lançamento '{number} - {seller}' é novo. Adicionando.")
                dados_para_adicionar.append(df_processado)
                arquivos_processados += 1
            else:
                log(f"AVISO: Lançamento '{number} - {seller}' já existe e foi ignorado.")
        
        if dados_para_adicionar:
            df_final = pd.concat([df_final] + dados_para_adicionar, ignore_index=True)
            log("Ordenando os dados pela coluna 'NUMBER'...")
            df_final = df_final.sort_values(by='NUMBER', ascending=True, kind='mergesort', ignore_index=True)

    if arquivos_processados == 0:
        log("Nenhuma operação necessária.")
        return True, "Nenhuma operação necessária."

    for col in ['Estado', 'Estoque', 'Local']:
        if col not in df_final.columns:
            df_final[col] = ""

    if not modo_atualizacao and 'UNIQUE ID' in df_final.columns:
        df_final['UNIQUE ID'] = range(1, len(df_final) + 1)
    
    df_final = df_final.reindex(columns=colunas_mestra_originais)
    msg_final = f"Operação concluída para {arquivos_processados} arquivo(s)."
    log(f"\n{msg_final}")
    
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_path = f"{os.path.splitext(planilha_mestra)[0]}_backup_{timestamp}.xlsx"
        log(f"Criando backup em: {os.path.basename(backup_path)}")
        shutil.copy(planilha_mestra, backup_path)

        log("Iniciando gravação direta na aba 'database'...")
        workbook = load_workbook(planilha_mestra)

        if "database" in workbook.sheetnames:
            db_sheet_to_remove = workbook["database"]
            workbook.remove(db_sheet_to_remove)
            log("  -> Aba 'database' existente foi removida para atualização.")

        db_sheet = workbook.create_sheet("database", 0)
        log("  -> Nova aba 'database' criada.")
        
        for r in dataframe_to_rows(df_final, index=False, header=True):
            db_sheet.append(r)
        
        if "Planilha1" in workbook.sheetnames:
              planilha1_a_remover = workbook["Planilha1"]
              workbook.remove(planilha1_a_remover)
              log("  -> Aba 'Planilha1' de rascunho foi removida.")

        log("Salvando o arquivo Excel...")
        workbook.save(planilha_mestra)
        log(f"\n--- SUCESSO! A planilha foi atualizada diretamente na aba 'database'. ---")
        return True, msg_final
    except Exception as e:
        log(f"ERRO ao salvar os dados: {e} \n {traceback.format_exc()}")
        return False, "Falha ao salvar os dados."

# CLASSE PRINCIPAL DA APLICAÇÃO
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        ctk.set_appearance_mode("light")
        self.title("Consolidador de Planilhas CUCALA")
        self.geometry("1280x720")
        
        def load_users():
            with open(resource_path("data/users.json"), "r", encoding="utf-8") as f:
                all_users = json.load(f)
            return all_users.get("consolidador", {})

        self.VALID_CREDENTIALS = load_users()
        
        self.planilha_mestra_path = ""
        self.planilhas_origem_paths = []
        self.worker_thread = None
        self.log_queue = queue.Queue()
        self.current_user = None
        
        self.show_login_frame() 

    def show_login_frame(self):
        self.login_frame = ctk.CTkFrame(self)
        self.login_frame.pack(fill="both", expand=True)
        self.login_frame.grid_columnconfigure(0, weight=1)
        self.login_frame.grid_rowconfigure(6, weight=1)
        
        try:
            logo_path = resource_path("assets/logo.png")
            logo_image = ctk.CTkImage(Image.open(logo_path), size=(151, 151))
            logo_label = ctk.CTkLabel(self.login_frame, image=logo_image, text="")
            logo_label.grid(row=0, column=0, pady=(100, 20))
        except Exception:
            logo_label = ctk.CTkLabel(self.login_frame, text="CUCALA", font=ctk.CTkFont(size=32, weight="bold"))
            logo_label.grid(row=0, column=0, pady=(100, 20))

        user_entry = ctk.CTkEntry(self.login_frame, placeholder_text="Usuário", width=250)
        user_entry.grid(row=1, column=0, padx=30, pady=10)
        
        pass_entry = ctk.CTkEntry(self.login_frame, placeholder_text="Senha", show="*", width=250)
        pass_entry.grid(row=2, column=0, padx=30, pady=10)

        remember_me_var = ctk.StringVar()
        config = configparser.ConfigParser()
        try:
            config.read(CONFIG_FILE_PATH)
            saved_user = config.get('Login', 'username', fallback=None)
            if saved_user:
                user_entry.insert(0, saved_user)
                remember_me_var.set("on")
        except Exception as e:
            logging.info(f"Não foi possível ler o arquivo de configuração: {e}")

        pass_entry.bind("<Return>", lambda event: self.check_login(user_entry, pass_entry, remember_me_var, status_label))
        
        login_button = ctk.CTkButton(self.login_frame, text="Login", width=250, 
                                      command=lambda: self.check_login(user_entry, pass_entry, remember_me_var, status_label))
        login_button.grid(row=4, column=0, padx=30, pady=15, ipady=5)
        
        status_label = ctk.CTkLabel(self.login_frame, text="", text_color="red")
        status_label.grid(row=5, column=0, padx=30, pady=5)

    def check_login(self, user_entry, pass_entry, remember_me_var, status_label):
        user = user_entry.get().lower()
        password = pass_entry.get()
        if user in self.VALID_CREDENTIALS and self.VALID_CREDENTIALS[user] == password:
            config = configparser.ConfigParser()
            config.read(CONFIG_FILE_PATH)
            if 'Login' not in config.sections():
                config.add_section('Login')

            if remember_me_var.get() == "on":
                config.set('Login', 'username', user)
            else:
                if config.has_option('Login', 'username'):
                    config.remove_option('Login', 'username')
            
            with open(CONFIG_FILE_PATH, 'w') as configfile:
                config.write(configfile)

            self.current_user = user
            self.login_frame.destroy()
            self.setup_main_ui()
        else:
            status_label.configure(text="Usuário ou senha inválida.")
    
    def carregar_configuracoes(self):
        config = configparser.ConfigParser()
        try:
            config.read(CONFIG_FILE_PATH)
            path = config.get('Paths', 'planilha_mestra', fallback=None)
            if path and os.path.exists(path):
                self.planilha_mestra_path = path
                self.label_mestra_path.configure(text=f"Arquivo Mestre: {os.path.basename(path)}")
            else:
                self.label_mestra_path.configure(text="Planilha Mestra não definida. Clique para configurar.")
        except Exception as e:
            logging.error(f"Erro ao carregar configurações: {e}")
            self.label_mestra_path.configure(text="Erro ao carregar config. Clique para definir.")

    def definir_planilha_mestra(self):
        path = filedialog.askopenfilename(
            title="Selecione a Planilha Mestra",
            filetypes=[("Planilhas Excel", "*.xlsx")]
        )
        if path:
            self.planilha_mestra_path = path
            self.label_mestra_path.configure(text=f"Arquivo Mestre: {os.path.basename(path)}")
            
            config = configparser.ConfigParser()
            config.read(CONFIG_FILE_PATH)
            if 'Paths' not in config.sections():
                config.add_section('Paths')
            config.set('Paths', 'planilha_mestra', path)
            with open(CONFIG_FILE_PATH, 'w') as configfile:
                config.write(configfile)
            
            messagebox.showinfo("Sucesso", "O caminho da planilha mestra foi salvo com sucesso!")

    def setup_main_ui(self):
        self.grid_columnconfigure(0, weight=1); self.grid_columnconfigure(1, weight=2)
        self.grid_rowconfigure(1, weight=1)
        
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, columnspan=2, padx=20, pady=10, sticky="ew")
        header.grid_columnconfigure(0, weight=1)
        
        ctk.CTkLabel(header, text="Consolidador de Planilhas CUCALA", font=ctk.CTkFont(size=28, weight="bold")).grid(row=0, column=0, sticky="w")
        
        mestra_header_frame = ctk.CTkFrame(header)
        mestra_header_frame.grid(row=1, column=0, columnspan=2, pady=(10,0), sticky="ew")
        self.label_mestra_path = ctk.CTkLabel(mestra_header_frame, text="Carregando...", anchor="w", font=ctk.CTkFont(size=14))
        self.label_mestra_path.pack(side="left", padx=10, fill="x", expand=True)
        
        ctk.CTkButton(mestra_header_frame, text="Definir / Alterar Planilha Mestra", width=220, command=self.definir_planilha_mestra).pack(side="right", padx=10, pady=5)
        
        left_frame = ctk.CTkFrame(self)
        left_frame.grid(row=1, column=0, padx=(20, 10), pady=10, sticky="nsew")
        left_frame.grid_columnconfigure(0, weight=1)
        left_frame.grid_rowconfigure(1, weight=1)

        try:
            logo_path = resource_path("assets/logo.png")
            logo_image = ctk.CTkImage(Image.open(logo_path), size=(142, 142))
            ctk.CTkLabel(left_frame, image=logo_image, text="").grid(row=0, column=0, pady=10)
        except Exception:
            ctk.CTkLabel(left_frame, text="CUCALA", font=ctk.CTkFont(size=24, weight="bold")).grid(row=0, column=0, pady=20)
        
        origem_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        origem_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        origem_frame.grid_columnconfigure(0, weight=1)
        origem_frame.grid_rowconfigure(1, weight=1)
        
        ctk.CTkLabel(origem_frame, text="1. Selecione as Planilhas de Origem:", font=ctk.CTkFont(size=16)).grid(row=0, column=0, sticky="w")
        listbox_container = ctk.CTkFrame(origem_frame)
        listbox_container.grid(row=1, column=0, pady=5, sticky="nsew")
        listbox_container.grid_columnconfigure(0, weight=1)
        listbox_container.grid_rowconfigure(0, weight=1)
        
        self.listbox = tkinter.Listbox(listbox_container, background="#EBEBEB", foreground="black", selectbackground="#1F6AA5", borderwidth=0, highlightthickness=0, font=("Calibri", 12), selectmode=tkinter.EXTENDED)
        self.listbox.grid(row=0, column=0, padx=1, pady=1, sticky="nsew")
        
        reorder_frame = ctk.CTkFrame(origem_frame, fg_color="transparent")
        reorder_frame.grid(row=2, column=0, pady=5, sticky="ew")
        reorder_frame.grid_columnconfigure((0,1,2), weight=1)
        ctk.CTkButton(reorder_frame, text="Selecionar Arquivos", command=self.selecionar_origem).grid(row=0, column=0, padx=(0,5), sticky="ew")
        ctk.CTkButton(reorder_frame, text="↑", command=self.move_up, width=40).grid(row=0, column=1, padx=5)
        ctk.CTkButton(reorder_frame, text="↓", command=self.move_down, width=40).grid(row=0, column=2, padx=(5,0))
        
        action_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        action_frame.grid(row=2, column=0, padx=10, pady=(10,20), sticky="ew")
        action_frame.grid_columnconfigure(0, weight=65)
        action_frame.grid_columnconfigure(1, weight=35)
        self.add_button = ctk.CTkButton(action_frame, text="Adicionar Novos Dados", height=50, font=ctk.CTkFont(size=18, weight="bold"), command=lambda: self.executar(modo_atualizacao=False))
        self.add_button.grid(row=0, column=0, ipady=10, sticky="ew", padx=(0,10))
        self.update_button = ctk.CTkButton(action_frame, text="Atualizar\nLançamento", height=50, font=ctk.CTkFont(size=14, weight="bold"), fg_color="#4A4D50", hover_color="#5F6266", command=lambda: self.executar(modo_atualizacao=True))
        self.update_button.grid(row=0, column=1, ipady=10, sticky="ew")
        
        right_frame = ctk.CTkFrame(self)
        right_frame.grid(row=1, column=1, padx=(10, 20), pady=10, sticky="nsew")
        right_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(right_frame, text="Caixa de diálogo do sistema:", font=ctk.CTkFont(size=16)).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.log_textbox = ctk.CTkTextbox(right_frame, font=("Courier New", 12))
        self.log_textbox.grid(row=1, column=0, padx=10, pady=(0,10), sticky="nsew")
        
        footer_frame = ctk.CTkFrame(self, fg_color="transparent")
        footer_frame.grid(row=2, column=0, columnspan=2, padx=20, pady=10, sticky="ew")
        footer_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(footer_frame, text="Desenvolvido por Vinicios Reis para uso exclusivo CUCALA", text_color="gray", font=ctk.CTkFont(size=12)).grid(row=0, column=0, sticky="w")
        
        self.theme_switch = ctk.CTkSwitch(footer_frame, text="Light", command=self.change_theme)
        self.theme_switch.grid(row=0, column=1, sticky="e")
        
        self.carregar_configuracoes()
        
    def change_theme(self):
        is_dark = self.theme_switch.get() == 1
        new_mode = "dark" if is_dark else "light"
        ctk.set_appearance_mode(new_mode)
        
        if is_dark:
            self.theme_switch.configure(text="Dark")
            self.listbox.configure(background="#343638", foreground="white")
        else:
            self.theme_switch.configure(text="Light")
            self.listbox.configure(background="#EBEBEB", foreground="black")

    def selecionar_origem(self):
        paths = filedialog.askopenfilenames(title="Selecione as Planilhas de Origem", filetypes=[("Planilhas Excel", "*.xlsx")])
        if paths:
            self.planilhas_origem_paths = list(paths); self.listbox.delete(0, ctk.END)
            for path in self.planilhas_origem_paths: self.listbox.insert(ctk.END, os.path.basename(path))

    def move_up(self):
        selected_indices = self.listbox.curselection()
        if not selected_indices: return
        for i in selected_indices:
            if i > 0:
                self.planilhas_origem_paths.insert(i-1, self.planilhas_origem_paths.pop(i)); text = self.listbox.get(i)
                self.listbox.delete(i); self.listbox.insert(i-1, text); self.listbox.selection_set(i-1)
    
    def move_down(self):
        selected_indices = self.listbox.curselection()
        if not selected_indices: return
        for i in reversed(selected_indices):
            if i < self.listbox.size() - 1:
                self.planilhas_origem_paths.insert(i+1, self.planilhas_origem_paths.pop(i)); text = self.listbox.get(i)
                self.listbox.delete(i); self.listbox.insert(i+1, text); self.listbox.selection_set(i+1)

    def process_queue(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                if isinstance(msg, tuple) and msg[0] == "WORKER_FINISHED":
                    self.finalize_process(msg[1])
                    return
                self.log_textbox.insert(ctk.END, msg + '\n')
                self.log_textbox.see(ctk.END)
        except queue.Empty:
            pass
        if self.worker_thread and self.worker_thread.is_alive():
            self.after(100, self.process_queue)

    def executar(self, modo_atualizacao):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Aguarde", "Um processo já está em andamento.")
            return
            
        self.log_textbox.delete("1.0", ctk.END)
        
        if not self.planilha_mestra_path or not os.path.exists(self.planilha_mestra_path):
            messagebox.showerror("Erro de Validação", "O caminho da Planilha Mestra não foi definido ou o arquivo não foi encontrado. Por favor, configure o caminho correto antes de continuar.")
            return

        if not self.planilhas_origem_paths:
            messagebox.showerror("Erro de Validação", "Nenhuma planilha de origem foi selecionada.")
            return
        
        arquivos_para_processar = []
        if modo_atualizacao:
            selected_indices = self.listbox.curselection()
            if len(selected_indices) != 1:
                messagebox.showerror("Seleção Inválida", "Para atualizar, por favor, selecione apenas UM lançamento na lista.")
                return
            arquivos_para_processar = [self.planilhas_origem_paths[selected_indices[0]]]
        else:
            arquivos_para_processar = self.planilhas_origem_paths

        self.add_button.configure(state="disabled")
        self.update_button.configure(state="disabled")
        
        self.worker_thread = threading.Thread(target=self.run_consolidation_worker, args=(arquivos_para_processar, modo_atualizacao))
        self.worker_thread.start()
        self.after(100, self.process_queue)

    def run_consolidation_worker(self, arquivos_para_processar, modo_atualizacao):
        try:
            sucesso, msg_final = executar_logica_consolidacao(self.planilha_mestra_path, arquivos_para_processar, self.log_queue, self.current_user, modo_atualizacao)
            self.log_queue.put(("WORKER_FINISHED", (sucesso, msg_final)))
        except Exception as e:
            tb_str = traceback.format_exc()
            self.log_queue.put(f"ERRO CRÍTICO NO WORKER: {e}\n{tb_str}")
            self.log_queue.put(("WORKER_FINISHED", (False, f"Erro crítico: {e}")))
            
    def finalize_process(self, result):
        sucesso, msg_final = result
        self.add_button.configure(state="normal")
        self.update_button.configure(state="normal")
        if sucesso:
            messagebox.showinfo("Sucesso", f"Processo concluído!\n\n{msg_final}")
        else:
            messagebox.showerror("Falha na Execução", f"O processo falhou.\n\nDetalhe: {msg_final}")

# PONTO DE ENTRADA DO PROGRAMA
if __name__ == "__main__":
    app = App()
    app.mainloop()